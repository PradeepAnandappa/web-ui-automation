import random
import string
import time
from pathlib import Path

import openpyxl
from pytest_testrail.plugin import pytestrail
import allure
from Configuration.config import config
from src.PageObjects.BasePage.LeftNavigationPanelPage import LeftNavigationPanelPage
from src.PageObjects.Login_Logout.LoginPage import LoginPage
from src.PageObjects.Monitor.MonitorPage import MonitorPage
from src.PageObjects.Monitor.TrackingScheduler.TrackingScheduleUI.LocateScheduleMainView.AddEditTrackingSchedulePage import \
    AddEditTrackingSchedulePage
from src.PageObjects.Monitor.TrackingScheduler.TrackingScheduleUI.LocateScheduleMainView.TrackingSchedulerPage import TrackingSchedulerPage
from src.BaseFile.BaseTest import BaseTest
from conftest import setup_driver, getEnvironment, getBrowser
from src.PageObjects.Reports.RunReports.RunReportsMainView.ReportPage import ReportPage


class Test_LocateSchedule(BaseTest):
    @pytestrail.case('C33106')
    def test_LocateScheduleView(self):
        self.loginPage = LoginPage(self.driver)
        self.loginPage.user_login(config.AccoundID, config.UserName, config.Password)
        self.monitorPage = MonitorPage(self.driver)
        self.trackingSchedulerPage = TrackingSchedulerPage(self.driver)
        page_found = self.monitorPage.getPageTitle("Monitor")
        assert page_found
        self.monitorPage.click_TrackingSchedularTab()
        page_URL = self.monitorPage.getPageURL()
        Expected_URL = config.baseURL + "/monitor/locate-schedule"
        assert Expected_URL == page_URL
        self.trackingSchedulerPage.click_AddNew()
        self.addEditTrackingSchedulePage = AddEditTrackingSchedulePage(self.driver)
        flag = self.addEditTrackingSchedulePage.validate_AddNewTrackingScheduleModal()
        assert flag
        scheduleName = ''.join(random.choices(string.ascii_letters, k=7))
        self.addEditTrackingSchedulePage.addTrackingSchedule(scheduleName)
        self.trackingSchedulerPage.validate_SearchRecord(scheduleName)
        self.trackingSchedulerPage.click_ExportToExcel()
        time.sleep(5)
        file_path = str(Path.home()) + "\\PycharmProjects\\allGeo_webApp\\src\\Resources\\Download\\Locate Schedules.xlsx"

        # Define variable to load the dataframe
        dataframe = openpyxl.load_workbook(file_path)

        # Define variable to read sheet
        dataframe1 = dataframe.active

        print("row count: " + str(dataframe1.max_row))
        assert 1 < dataframe1.max_row

        # Iterate the loop to read the cell values
        for row in range(0, dataframe1.max_row):
            for col in dataframe1.iter_cols(1, dataframe1.max_column):
                print(col[row].value)

        self.trackingSchedulerPage.deleteFile(file_path)



