import time
from pathlib import Path

from pytest_testrail.plugin import pytestrail
from Configuration.config import config
from src.PageObjects.Login_Logout.LoginPage import LoginPage
from src.PageObjects.Monitor.MonitorPage import MonitorPage
from src.PageObjects.BasePage.LeftNavigationPanelPage import LeftNavigationPanelPage
from src.PageObjects.Company.Employees.MainView.EmployeesCRUD_UI.EmployeesPage import EmployeesPage
from src.BaseFile.BaseTest import BaseTest
import openpyxl


class Test_Company(BaseTest):
    @pytestrail.case('C34794')
    def test_addEmployee(self):
        self.loginPage = LoginPage(self.driver)
        self.loginPage.user_login(config.AccoundID, config.UserName, config.Password)
        self.monitorPage = MonitorPage(self.driver)
        page_found = self.monitorPage.getPageTitle("Monitor")
        assert page_found
        self.leftNavigationPanelPage = LeftNavigationPanelPage(self.driver)
        self.leftNavigationPanelPage.click_Company()
        self.employeesPage = EmployeesPage(self.driver)
        time.sleep(5)
        page_found = self.employeesPage.validate_page_title()
        assert page_found
        self.employeesPage.select_ThreeDotMenuColumnsOption("Employee ID")

        self.employeesPage.click_ExportToExcel()
        time.sleep(5)
        file_path = str(Path.home()) + "\\PycharmProjects\\allGeo_webApp\\src\\Resources\\Download\\Employees.xlsx"

        # Define variable to load the dataframe
        dataframe = openpyxl.load_workbook(file_path)

        self.employeesPage.deleteFile(file_path)

        # Define variable to read sheet
        dataframe1 = dataframe.active

        print("row count: "+str(dataframe1.max_row))
        assert 1 < dataframe1.max_row

        # Iterate the loop to read the cell values
        for row in range(0, 1):
            for col in dataframe1.iter_cols(1, dataframe1.max_column):
                print(col[row].value)
                if "Employee ID" == col[row].value:
                    assert False
                    break

        self.employeesPage.reset_ThreeDotSelection()





